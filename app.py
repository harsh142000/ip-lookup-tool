from flask import Flask, request, jsonify, send_file, render_template
import os
import requests
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from iso3166 import countries
import ipaddress
import tempfile
import socket
from urllib.parse import urlparse
import time
import io
import io
from flask import Flask, request, jsonify, send_file, render_template
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

load_dotenv()

app = Flask(__name__)

VT_KEYS = [key.strip() for key in os.getenv("VT_API_KEYS", "").split(",") if key.strip()]
ABUSEIPDB_KEY = os.getenv("ABUSEIPDB_API_KEY")
DBIP_KEY = os.getenv("DBIP_API_KEY")
IPINFO_KEY = os.getenv("IPINFO_API_KEY")
APIVOID_KEY = os.getenv("APIVOID_API_KEY")

vt_key_index = 0
vt_key_lock = threading.Lock()
exhausted_vt_keys = set()
exhausted_other_keys = set()
vt_keys_used = set()
vt_keys_success = set()

MAX_WORKERS = 100

used_services = set()
unused_services = set()
country_cache = {}

def get_country_name(code):
    if not code:
        return "Unknown"
    if code in country_cache:
        return country_cache[code]
    try:
        name = countries.get(code.upper()).name
        country_cache[code] = name
        return name
    except:
        return code

def mask_key(key):
    return key[:4] + "..." + key[-4:] if key else "None"

def get_next_vt_key():
    global vt_key_index
    with vt_key_lock:
        for _ in range(len(VT_KEYS)):
            key = VT_KEYS[vt_key_index % len(VT_KEYS)]
            vt_key_index += 1
            if key and key not in exhausted_vt_keys:
                return key
        return None

import base64

import base64

def fetch_virustotal_url_data(url):
    result = {
        "detections": None,
        "categories": [],
        "vt_key_used": None,
        "status_codes": {},       # ← new
        "services_used": []
    }

    key = get_next_vt_key()
    if not key:
        return result

    headers = {"x-apikey": key}
    try:
        # submit URL
        post_resp = requests.post(
            "https://www.virustotal.com/api/v3/urls",
            headers=headers,
            data={"url": url},
            timeout=10
        )
        result["status_codes"]["VT_URL_submit"] = post_resp.status_code

        if post_resp.status_code != 200:
            exhausted_vt_keys.add(key)
            return result

        # lookup report
        encoded = base64.urlsafe_b64encode(url.encode()).decode().rstrip("=")
        get_resp = requests.get(
            f"https://www.virustotal.com/api/v3/urls/{encoded}",
            headers=headers,
            timeout=10
        )
        result["status_codes"]["VT_URL_report"] = get_resp.status_code

        if get_resp.status_code != 200:
            exhausted_vt_keys.add(key)
            return result

        data = get_resp.json().get("data", {}).get("attributes", {})
        result["detections"] = data.get("last_analysis_stats", {}).get("malicious", 0)
        result["categories"] = list(data.get("categories", {}).values())
        result["vt_key_used"] = key
        result["services_used"].append("VirusTotal URL")

        vt_keys_used.add(key)
        vt_keys_success.add(key)

    except Exception as e:
        print(f"[ERROR] VT URL scan failed: {e}")
        # you may also set result["status_codes"]["VT_URL_error"] = str(e)

    return result


# -------------------------------
# VirusTotal IP Query with Key Rotation
# -------------------------------
def query_virustotal(ip):
    tried = set()
    while True:
        key = get_next_vt_key()
        if not key or key in tried:
            break
        
        tried.add(key)
    
        headers = {"x-apikey": key}
        vt_keys_used.add(key)
        try:
            resp = requests.get(f"https://www.virustotal.com/api/v3/ip_addresses/{ip}", headers=headers, timeout=10)
            if resp.status_code in (401, 403):
                exhausted_vt_keys.add(key)
                continue
            if resp.status_code != 200:
                return {}, f"VTError {resp.status_code}", key
            data = resp.json().get("data", {}).get("attributes", {})
            vt_keys_success.add(key)
            return {
                "isp": data.get("as_owner"),
                "country": get_country_name(data.get("country")),
                "detections": data.get("last_analysis_stats", {}).get("malicious", 0)
            }, "VT", key
        except Exception as e:
            exhausted_vt_keys.add(key)
            return {}, f"VT Exception: {str(e)}", key
    return {}, "NoVTKeyAvailable", None


# -------------------------------
# AbuseIPDB Query with Rate-Limit Handling
# -------------------------------
def query_abuseipdb(ip):
    if not ABUSEIPDB_KEY:
        return {}, "NoKey"
    headers = {"Key": ABUSEIPDB_KEY, "Accept": "application/json"}
    try:
        resp = requests.get(f"https://api.abuseipdb.com/api/v2/check?ipAddress={ip}&maxAgeInDays=90", headers=headers, timeout=10)
        if resp.status_code == 429:
            exhausted_other_keys.add("AbuseIPDB")
            return {}, "RateLimit"
        if resp.status_code != 200:
            return {}, f"AbuseIPDB Error {resp.status_code}"
        data = resp.json().get("data", {})
        used_services.add("AbuseIPDB")
        return {
            "isp": data.get("isp"),
            "country": get_country_name(data.get("countryCode")),
            "detections": data.get("totalReports", 0)
        }, "AbuseIPDB"
    except Exception as e:
        return {}, f"AbuseIPDB Exception: {str(e)}"


# -------------------------------
# DB-IP Query (No Rate-Limit)
# -------------------------------
def query_dbip(ip):
    if not DBIP_KEY:
        return {}, "NoKey"
    try:
        resp = requests.get(f"https://api.db-ip.com/v2/{DBIP_KEY}/{ip}/asn.json", timeout=10)
        if resp.status_code != 200:
            return {}, f"DBIP Error {resp.status_code}"
        data = resp.json()
        used_services.add("DBIP")
        return {
            "isp": data.get("organisation"),
            "country": get_country_name(data.get("countryCode")),
            "detections": 0
        }, "DBIP"
    except Exception as e:
        return {}, f"DBIP Exception: {str(e)}"


# -------------------------------
# IPInfo Query (No Rate-Limit)
# -------------------------------
def query_ipinfo(ip):
    try:
        headers = {"Authorization": f"Bearer {IPINFO_KEY}"} if IPINFO_KEY else {}
        resp = requests.get(f"https://ipinfo.io/{ip}/json", headers=headers, timeout=10)
        if resp.status_code != 200:
            return {}, f"IPINFO Error {resp.status_code}"
        data = resp.json()
        used_services.add("IPINFO")
        return {
            "isp": data.get("org"),
            "country": get_country_name(data.get("country")),
            "detections": 0
        }, "IPINFO"
    except Exception as e:
        return {}, f"IPINFO Exception: {str(e)}"


# -------------------------------
# APIVoid Query with Rate-Limit Handling
# -------------------------------
def query_apivoid(ip):
    if not APIVOID_KEY:
        return {}, "NoKey"
    try:
        resp = requests.get(f"https://endpoint.apivoid.com/iprep/v1/pay-as-you-go/?key={APIVOID_KEY}&ip={ip}", timeout=10)
        if resp.status_code == 429:
            exhausted_other_keys.add("APIVoid")
            return {}, "RateLimit"
        if resp.status_code != 200:
            return {}, f"APIVoid Error {resp.status_code}"
        data = resp.json().get("data", {}).get("report", {})
        used_services.add("APIVoid")
        return {
            "isp": data.get("network", {}).get("organization"),
            "country": get_country_name(data.get("information", {}).get("country_code")),
            "detections": data.get("blacklists", {}).get("detections", 0)
        }, "APIVoid"
    except Exception as e:
        return {}, f"APIVoid Exception: {str(e)}"
        
def is_valid_public_ip(ip):
    try:
        ip_obj = ipaddress.ip_address(ip)
        return not (ip_obj.is_private or ip_obj.is_loopback or ip_obj.is_unspecified)
    except:
        return False

import re
from urllib.parse import urlparse

def is_valid_url(url):
    try:
        parsed = urlparse(url if url.startswith("http") else f"http://{url}")
        hostname = parsed.hostname

        if not hostname or '.' not in hostname:
            return False

        # Reject short numeric-like hostnames (e.g., "12.23.4")
        if re.fullmatch(r"\d{1,3}(\.\d{1,3}){1,2}", hostname):
            return False

        # Check TLD
        if not re.search(r"\.[a-zA-Z]{2,}$", hostname):
            return False

        return True
    except:
        return False

import re

def is_valid_ip(ip):
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        return False

def get_ip_info(ip):
    ip_info = {
        "ip": ip,
        "asn": "",
        "isp": "",
        "country": "",
        "detections": 0,
        "used_service": "",
        "used_key": "",
        "status_codes": {},
        # Track which service gave which piece of data
        "service_sources": {
            "asn":    None,
            "isp":    None,
            "country":None,
            "detections": None
        }
    }

    # === VirusTotal Primary Lookup ===
    for _ in range(len(VT_KEYS)):
        vt_key = get_next_vt_key()
        if not vt_key:
            break

        used_services.add("VirusTotal")
        headers = {"x-apikey": vt_key}
        try:
            resp = requests.get(
                f"https://www.virustotal.com/api/v3/ip_addresses/{ip}",
                headers=headers,
                timeout=10
            )
            ip_info["status_codes"]["VirusTotal"] = resp.status_code
            vt_keys_used.add(vt_key)

            if resp.status_code == 200:
                data    = resp.json().get("data", {}).get("attributes", {})
                asn_vt  = data.get("asn", "")
                isp_vt  = data.get("as_owner", "")
                ctr_vt  = data.get("country", "")
                det_vt  = data.get("last_analysis_stats", {}).get("malicious", 0) or 0

                # Always take detections from VT
                ip_info["detections"] = det_vt
                ip_info["service_sources"]["detections"] = "VirusTotal"

                # Fill ASN/ISP/Country only if VT provided them
                if asn_vt:
                    ip_info["asn"] = asn_vt
                    ip_info["service_sources"]["asn"] = "VirusTotal"
                if isp_vt:
                    ip_info["isp"] = isp_vt
                    ip_info["service_sources"]["isp"] = "VirusTotal"
                if ctr_vt:
                    ip_info["country"] = get_country_name(ctr_vt)
                    ip_info["service_sources"]["country"] = "VirusTotal"

                ip_info["used_service"] = "VirusTotal"
                ip_info["used_key"]     = vt_key
                vt_keys_success.add(vt_key)

                # break early if ISP or Country came from VT
                if isp_vt or ctr_vt:
                    break

            elif resp.status_code in (401, 403):
                exhausted_vt_keys.add(vt_key)

        except Exception:
            ip_info["status_codes"]["VirusTotal"] = "Error"
            exhausted_vt_keys.add(vt_key)


    # === AbuseIPDB Fallback ===
    if not ip_info["isp"] and ABUSEIPDB_KEY:
        try:
            resp = requests.get(
                "https://api.abuseipdb.com/api/v2/check",
                headers={"Key": ABUSEIPDB_KEY, "Accept": "application/json"},
                params={"ipAddress": ip, "maxAgeInDays": "90"},
                timeout=10
            )
            ip_info["status_codes"]["AbuseIPDB"] = resp.status_code

            if resp.status_code == 200:
                data = resp.json().get("data", {})
                if data.get("asn"):
                    ip_info["asn"] = data["asn"]
                    ip_info["service_sources"]["asn"] = "AbuseIPDB"
                ip_info["isp"] = data.get("isp", "")
                ip_info["service_sources"]["isp"] = "AbuseIPDB"
                ip_info["country"] = get_country_name(data.get("countryCode", ""))
                ip_info["service_sources"]["country"] = "AbuseIPDB"
                ip_info["used_service"] = "AbuseIPDB"
                ip_info["used_key"]     = ABUSEIPDB_KEY
                used_services.add("AbuseIPDB")

        except Exception:
            ip_info["status_codes"]["AbuseIPDB"] = "Error"

    # === DB-IP Fallback ===
    if not ip_info["isp"] and DBIP_KEY:
        try:
            resp = requests.get(f"https://api.db-ip.com/v2/{DBIP_KEY}/{ip}/asn", timeout=10)
            ip_info["status_codes"]["DBIP"] = resp.status_code

            if resp.status_code == 200:
                data = resp.json()
                if data.get("asn"):
                    ip_info["asn"] = data["asn"]
                    ip_info["service_sources"]["asn"] = "DBIP"
                if data.get("organization"):
                    ip_info["isp"] = data["organization"]
                    ip_info["service_sources"]["isp"] = "DBIP"
                ip_info["country"] = get_country_name(data.get("countryCode", ""))
                ip_info["service_sources"]["country"] = "DBIP"
                ip_info["used_service"] = "DBIP"
                ip_info["used_key"]     = DBIP_KEY
                used_services.add("DBIP")

        except Exception:
            ip_info["status_codes"]["DBIP"] = "Error"

    # === IPInfo Fallback ===
    if not ip_info["isp"] and IPINFO_KEY:
        try:
            resp = requests.get(f"https://ipinfo.io/{ip}/json?token={IPINFO_KEY}", timeout=10)
            ip_info["status_codes"]["IPInfo"] = resp.status_code

            if resp.status_code == 200:
                data = resp.json()
                if data.get("org"):
                    ip_info["isp"] = data["org"]
                    ip_info["service_sources"]["isp"] = "IPInfo"
                ip_info["country"] = get_country_name(data.get("country", ""))
                ip_info["service_sources"]["country"] = "IPInfo"
                ip_info["used_service"] = "IPInfo"
                ip_info["used_key"]     = IPINFO_KEY
                used_services.add("IPINFO")

        except Exception:
            ip_info["status_codes"]["IPInfo"] = "Error"

    # === APIVoid Fallback ===
    if not ip_info["isp"] and APIVOID_KEY:
        try:
            resp = requests.get(
                f"https://endpoint.apivoid.com/iprep/v1/pay-as-you-go/?key={APIVOID_KEY}&ip={ip}",
                timeout=10
            )
            ip_info["status_codes"]["APIVoid"] = resp.status_code

            if resp.status_code == 200:
                rpt    = resp.json().get("data", {}).get("report", {})
                server = rpt.get("server", {})
                if server.get("isp"):
                    ip_info["isp"] = server["isp"]
                    ip_info["service_sources"]["isp"] = "APIVoid"
                ip_info["country"] = get_country_name(server.get("country_code", ""))
                ip_info["service_sources"]["country"] = "APIVoid"
                ip_info["used_service"] = "APIVoid"
                ip_info["used_key"]     = APIVOID_KEY
                used_services.add("APIVoid")

        except Exception:
            ip_info["status_codes"]["APIVoid"] = "Error"
    ip_info["summary"] = (
        f"The IP: {ip_info['ip']} belongs to ISP: {ip_info['isp'] or 'N/A'}, "
        f"from Country: {ip_info['country'] or 'N/A'}, with Detection count: {ip_info['detections']}"
    )
    return ip_info


def lookup_url(url):
    vt_data = fetch_virustotal_url_data(url)
    used_services.add("VT")

    # Always take detections from VT (0 if missing)
    detections = vt_data.get("detections") or 0

    return {
        "type": "URL",
        "query": url,
        "hostname": urlparse(url if url.startswith("http") else f"http://{url}").hostname,
        "resolved_ip": "-",         # No IP resolution
        "ip": url,
        "asn": "",                  # N/A for URLs
        "isp": "N/A",
        "country": "N/A",
        "detections": detections,
        "vt_key_used": mask_key(vt_data.get("vt_key_used")) if vt_data.get("vt_key_used") else None,

        # Track exactly which service supplied each field
        "service_sources": {
            "asn":        None,
            "isp":        None,
            "country":    None,
            "detections": "VirusTotal URL"
        },

        # HTTP status codes from VT URL submit & report
        "status_codes": vt_data.get("status_codes", {}),

        "summary": (
            f"The URL: {url} was scanned by VirusTotal with {detections} malicious detections."
            + (f" Categories: {', '.join(vt_data.get('categories', []))}." if vt_data.get('categories') else "")
        )
    }
    vt_data = fetch_virustotal_url_data(url)
    used_services.add("VT")
    detections = vt_data.get("detections") or 0
    categories = vt_data.get("categories", [])
    vt_key_used = vt_data.get("vt_key_used")

    return {
        "type": "URL",
        "query": url,
        "hostname": urlparse(url if url.startswith("http") else f"http://{url}").hostname,
        "resolved_ip": "-",  # No IP resolution anymore
        "ip": url,
        "isp": "N/A",        # Skipping ISP/Country enrichment
        "country": "N/A",
        "detections": detections,
        "vt_key_used": mask_key(vt_key_used) if vt_key_used else None,
        
        "summary": (
            f"The URL: {url} was found in VirusTotal with {detections} malicious detections."
            + (f" Categories: {', '.join(categories)}." if categories else "")
        )
    }

# ✅ `handle_ip_lookup()` and `/download_excel` + `/` route are included in [next message] due to length...
@app.route("/get_ip_info", methods=["POST"])
def handle_ip_lookup():
    start = time.time()
    data = request.json
    entries = data.get("ips", [])

    global used_services, unused_services, vt_keys_used, vt_keys_success, exhausted_other_keys
    used_services.clear()
    unused_services.clear()
    vt_keys_used.clear()
    vt_keys_success.clear()
    exhausted_other_keys.clear()

    seen = set()
    valid_entries = []
    for entry in entries:
        e = entry.strip()
        if e in seen: continue
        seen.add(e)
        if is_valid_public_ip(e) or is_valid_url(e):
            valid_entries.append(e)
        if len(valid_entries) >= 100: break

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        results = list(executor.map(
            lambda e: get_ip_info(e) if is_valid_public_ip(e) else lookup_url(e),
            valid_entries
        ))

    has_url = any(r.get("type") == "URL" for r in results)

    no_data_ips = []
    for r in results:
        is_url      = (r.get("type") == "URL")
        det         = r.get("detections", 0)
        isp, ctr    = r.get("isp",""), r.get("country","")
        if is_url:
            if det == 0 and not isp and not ctr:
                no_data_ips.append(r["ip"])
        else:
            if not isp and not ctr:
                no_data_ips.append(r["ip"])

    table_rows = ""
    raw_table  = []
    for r in results:
        ip_or_url = r.get("query", r["ip"])
        resolved  = r.get("resolved_ip", "-")
        isp, ctr  = r["isp"], r["country"]
        det       = r["detections"]
        table_rows += (
            f"<tr>"
            f"<td>{ip_or_url}</td>"
            f"<td>{resolved}</td>"
            f"<td>{isp}</td>"
            f"<td>{ctr}</td>"
            f"<td>{det}</td>"
            f"</tr>"
        )
        raw_table.append([ip_or_url, resolved, isp, ctr, det, r.get("used_key")])

    summary_lines = []
    for i, r in enumerate(results):
        if i: summary_lines.append("")
        summary_lines.append(r["summary"])
    summary_text = "\n".join(summary_lines)

    elapsed = round(time.time() - start, 2)
    unused_services.update({"VT","AbuseIPDB","DBIP","IPINFO","APIVoid"} - used_services)

    vt_ok   = vt_keys_used & vt_keys_success
    vt_bad  = exhausted_vt_keys.copy()

    # ← your original‐style summary prints
    print("\n📊 API USAGE SUMMARY")
    print(f"✅ Data found for {len(valid_entries)} entr{'y' if len(valid_entries) == 1 else 'ies'} in {elapsed} seconds.")
    print(f"🔧 Services Used     : {', '.join(sorted(used_services)) or 'None'}")
    print(f"⚪ Services Unused   : {', '.join(sorted(unused_services)) or 'None'}")
    print(f"✅ Successfully Used VT Keys: {len(vt_ok)}")
    for k in vt_ok:           print(f"    {mask_key(k)}")
    print(f"❌ Exhausted VT Keys: {len(vt_bad)}")
    for k in vt_bad:          print(f"    {mask_key(k)}")
    if exhausted_other_keys:
        print("❌ Exhausted Other Services:", ", ".join(exhausted_other_keys))
    if len(vt_bad) > 10:
        print("⚠️ Warning: More than 10 VT keys are exhausted. Consider rotating or refreshing your keys.")
    vt_unused = set(VT_KEYS) - (vt_keys_success | exhausted_vt_keys)
    print(f"🟡 Unused VT Keys: {len(vt_unused)}")
    for k in vt_unused:       print(f"    {mask_key(k)}")
    print("Used API Keys:")
    if vt_ok:    print("  VT Keys:", ", ".join(mask_key(k) for k in vt_ok))
    if "AbuseIPDB" in used_services: print("  AbuseIPDB Key:", mask_key(ABUSEIPDB_KEY))
    if "DBIP" in used_services:      print("  DBIP Key:", mask_key(DBIP_KEY))
    if "IPINFO" in used_services:    print("  IPInfo Key:", mask_key(IPINFO_KEY))
    if "APIVoid" in used_services:   print("  APIVoid Key:", mask_key(APIVOID_KEY))

    # ← per‐entry with service tags
      # Per-entry detailed printout
        # Per-entry single‑line detailed printout
        # Per‑entry concise but readable summary
        # Per‑entry summary with spacing
    print("\n📋 Per Entry Summary:\n")
    for r in results:
        ip_entry = r.get("ip") or r.get("query")
        if not ip_entry:
            continue

        # Compose the main fields with their sources
        main_parts = [
            f"ASN: {r.get('asn','N/A')} (from {r['service_sources'].get('asn') or '—'})",
            f"ISP: {r.get('isp','N/A')} (from {r['service_sources'].get('isp') or '—'})",
            f"Country: {r.get('country','N/A')} (from {r['service_sources'].get('country') or '—'})",
            f"Detections: {r.get('detections',0)} (from {r['service_sources'].get('detections') or '—'})\n"
        ]

        # Compose status codes
        status_parts = [f"{svc}={code}" for svc, code in r.get("status_codes", {}).items()]

        # Blank line for spacing, then the summary line
        print()
        print(
            f"[{ip_entry}]  " +
            ";   ".join(main_parts) +
            "\n|  StatusCodes: " + ", ".join(status_parts)
        )
        # Extra blank line after each, if you want even more
        print()

    print("----------------------------\n")

    types = { "IP" if is_valid_ip(e) else "URL" for e in entries if is_valid_ip(e) or is_valid_url(e) }
    column_label = "IP" if types=={"IP"} else "URL" if types=={"URL"} else "IP/URL"

    return jsonify({
        "summary":       summary_text,
        "table":         table_rows,
        "raw_table":     raw_table,
        "no_data_ips":   no_data_ips,
        "services_used":  sorted(used_services),
        "per_ip_vt_keys": {
            r["ip"]: {
                "used_service": r.get("used_service"),
                "used_key":     r.get("used_key"),
                "status_codes": r.get("status_codes",{})
            }
            for r in results if "ip" in r
        },
        "has_url":       has_url,
        "column_label":  column_label
    })

from flask import request, send_file
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

@app.route("/download_excel", methods=["POST"])
def download_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    data = request.get_json()
    table_data = data.get("table_data", [])
    summary_text = data.get("summary", "")
    column_label = data.get("column_label", "IP")

    print("Incoming /download_excel payload:")
    print("Summary:", summary_text.strip())
    print("Column label:", column_label)
    if table_data:
        print("First row of table_data:", table_data[0])

    wb = Workbook()
    ws_table = wb.active
    ws_table.title = "Lookup Data"

    # Determine if "Resolved IP" column is present
    has_resolved_ip = any(
        isinstance(row, list) and len(row) >= 5 and str(row[1]).strip() not in ("-", "", "None")
        for row in table_data
    )


    # Header
    headers = (
        [column_label, "Resolved IP", "ISP", "Country", "Detections"]
        if has_resolved_ip
        else [column_label, "ISP", "Country", "Detections"]
    )
    ws_table.append(headers)

    # Body
    for row in table_data:
        ip_or_url = row[0]
        resolved_ip = row[1]
        isp = row[2]
        country = row[3]
        detections = row[4]

        if resolved_ip and resolved_ip not in ("-", "", "None") and resolved_ip != ip_or_url:
            display_value = f"{ip_or_url}"
        else:
            display_value = ip_or_url

        if has_resolved_ip:
            ws_table.append([display_value, resolved_ip, isp, country, detections])
        else:
            ws_table.append([display_value, isp, country, detections])


    # Formatting styles
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Apply formatting
    for row in ws_table.iter_rows():
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border
            if cell.row == 1:
                cell.font = bold_font

    # Autofit column width
    for col_cells in ws_table.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws_table.column_dimensions[col_letter].width = max(12, min(max_length + 4, 50))

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Scan Summary"
    ws_summary["A1"].font = Font(size=14, bold=True)
    ws_summary["A2"] = summary_text.strip()
    ws_summary["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_summary.column_dimensions["A"].width = 100

    # Return Excel file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="IP_Info.xlsx"
    )

@app.route("/")
def index():
    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)
